import openpyxl
import re


workbook = openpyxl.load_workbook('not in preferred name.xlsx')
worksheet = workbook.worksheets[0]
max_row = worksheet.max_row
synonyms, conceptId = [], []
for x in range(2, max_row + 1):
    conceptId.append(worksheet.cell(row=x, column=1).value)
    synonyms.append(worksheet.cell(row=x, column=2).value)

# print(conceptId)
print("Phase 1")
# dup_Id, dup_num = [], []
dic = {}
workbook1 = openpyxl.load_workbook('duplicate.xlsx')
worksheet1 = workbook1.worksheets[0]
max_row1 = worksheet1.max_row
for x in range(2, max_row1 + 1):
    key = worksheet1.cell(row=x, column=1).value
    value = worksheet1.cell(row=x, column=2).value
    dic[key] = value

# Not Duplicate conceptId
concept_Id = []
workbook2 = openpyxl.load_workbook('result.xlsx')
worksheet2 = workbook2.worksheets[0]
max_row2 = worksheet2.max_row
for x in range(2, max_row2 + 1):
    concept_Id.append(worksheet2.cell(row=x, column=1).value)
# print(concept_Id)
# print(len(concept_Id))
print("Phase 2")
# for i, v in enumerate(conceptId):
#     if i % 1000 == 0:
#         print(i)
#     if v in dup_Id:
#         p = dup_Id.index(v)
#         num = dup_num[p]
#         string = Map[i]
#         for j in range(1, num + 1):
#             string0 = Map[i+j]
#             string = string + "\n" + string0

worksheet2.cell(row=1, column=2).value = "Synonyms"
synonyms_new = []
for i, v in enumerate(concept_Id):
    if i % 1000 == 0:
        print(i)
    try:
        num = dic.get(v)
        if num == 1:
            p = conceptId.index(v)
            synonyms_new.append(synonyms[p])
        elif num is not None:
            p = conceptId.index(v)
            string = synonyms[p]
            a = 0
            for j in range(2, num + 1):
                a += 1
                string0 = synonyms[p + a]
                string = str(string) + "\n" + str(string0)
            synonyms_new.append(string)
        else:
            string = ''
            synonyms_new.append(string)
    except ValueError:
        string = ''
        synonyms_new.append(string)

for i, v in enumerate(synonyms_new):
    worksheet2.cell(row=i + 2, column=2).value = synonyms_new[i]

workbook2.save('result.xlsx')
