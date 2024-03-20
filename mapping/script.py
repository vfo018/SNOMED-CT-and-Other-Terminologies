import openpyxl
import re



workbook = openpyxl.load_workbook('mapping.xlsx')
worksheet = workbook.worksheets[0]
max_row = worksheet.max_row
term, conceptId, Map = [], [], []
for x in range(2, max_row + 1):
    conceptId.append(worksheet.cell(row=x, column=1).value)
    Map.append(worksheet.cell(row=x, column=3).value)

# print(conceptId)
print("Phase 1")
# dup_Id, dup_num = [], []
dic = {}
workbook1 = openpyxl.load_workbook('mapping_duplicate.xlsx')
worksheet1 = workbook1.worksheets[0]
max_row1 = worksheet1.max_row
for x in range(2, max_row1 + 1):
    key = worksheet1.cell(row=x, column=1).value
    value = worksheet1.cell(row=x, column=2).value
    dic[key] = value

# Not Duplicate conceptId
concept_Id = []
workbook2 = openpyxl.load_workbook('result.xlsx')
worksheet2 = workbook2.worksheets[0]
max_row2 = worksheet2.max_row
for x in range(2, max_row2 + 1):
    concept_Id.append(worksheet2.cell(row=x, column=1).value)
# print(concept_Id)
# print(len(concept_Id))
print("Phase 2")
# for i, v in enumerate(conceptId):
#     if i % 1000 == 0:
#         print(i)
#     if v in dup_Id:
#         p = dup_Id.index(v)
#         num = dup_num[p]
#         string = Map[i]
#         for j in range(1, num + 1):
#             string0 = Map[i+j]
#             string = string + "\n" + string0

worksheet2.cell(row=1, column=2).value = "Default ICD-10 Mapping"
Map_new = []
for i, v in enumerate(concept_Id):
    if i % 1000 == 0:
        print(i)
    try:
        num = dic.get(v)
        if num == 1:
            p = conceptId.index(v)
            Map_new.append(Map[p])
        elif num is not None:
            p = conceptId.index(v)
            string = Map[p]
            a = 0
            for j in range(2, num + 1):
                a += 1
                string0 = Map[p + a]
                string = string + "^" + string0
            Map_new.append(string)
        else:
            Map_new.append('#NC')
    except ValueError:
        # 如果找不到mapping
        string = '#NC'
        Map_new.append(string)

for i, v in enumerate(Map_new):
    worksheet2.cell(row=i + 2, column=2).value = Map_new[i]

workbook2.save('result.xlsx')
