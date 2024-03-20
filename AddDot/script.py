import openpyxl
import re


workbook1 = openpyxl.load_workbook('mapping.xlsx')
worksheet1 = workbook1.worksheets[0]
max_row1 = worksheet1.max_row
ICD10_Nodots = []
for x in range(2, max_row1 + 1):
    ICD10_Nodots.append(worksheet1.cell(row=x, column=2).value)

workbook2 = openpyxl.load_workbook('ICD10_Edition5_CodesAndTitlesAndMetadata_GB_20160401.xlsx')
worksheet2 = workbook2.worksheets[0]
max_row2 = worksheet2.max_row
dic = {}
for x in range(2, max_row2 + 1):
    key = worksheet2.cell(row=x, column=2).value
    value = worksheet2.cell(row=x, column=1).value
    dic[key] = value

ICD10_Dots = []
for i, v in enumerate(ICD10_Nodots):
    if v != '#NC' or v != 'NEW' or v != '#HLT' or v != '#EPO':
        # Whether the code ends up with A or D
        ret = re.match(r".*[A|D]$", v)
        if ret:
            #  If yes, drop the last character
            v = v[:-1]
        v_dots = dic.get(v)
        if v_dots is not None:
            ICD10_Dots.append(v_dots)
        else:
            ICD10_Dots.append(v)
    else:
        ICD10_Dots.append(v)

# print(dic)
# print(ICD10_Dots)
worksheet1.cell(row=1, column=3).value = 'MapTargetDots'
for i, v in enumerate(ICD10_Dots):
    worksheet1.cell(row=i + 2, column=3).value = v

workbook1.save('mapping.xlsx')
